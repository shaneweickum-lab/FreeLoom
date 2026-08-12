"""
Converts freeloom_academic_database.xlsx (output of ml/scripts/freeloom_scraper.py)
into an `insert into research_citations` SQL file.

Usage:
    pip install openpyxl
    python3 convert_research_spreadsheet.py freeloom_academic_database.xlsx research_citations_import.sql

Column mapping (source column -> research_citations column):
    Activity / Methodology            -> title
    Category                          -> category
    Suggested Course Title            -> topic
    Primary Subject                   -> primary_subject
    Secondary Subject                 -> secondary_subject
    Description / Educational Rationale -> summary
    Classification Keywords           -> keywords (comma-split into text[])
    Scientific Backing (Source)       -> source
    Evidence Level                    -> evidence_level
    DOI / URL                         -> source_url

Rows missing a title, category, or summary are skipped (can't insert a
not-null column with nothing to put there) and counted in the stats report.
Duplicate titles within the sheet are kept -- ERIC and Crossref cover
different, non-overlapping records, and freeloom_scraper.py's own
drop_duplicates() already ran before this file was produced.

Also drops any row touching gender/sex/sexuality/religion topics (checked
across every text field, not just the title) -- FreeLoom's Research Library
is scoped to pedagogy/methodology evidence, deliberately excluding those
subject areas regardless of how a citation got pulled in by the scraper's
broad keyword search.
"""

import re
import sys
from openpyxl import load_workbook

CHUNK_SIZE = 100

EXCLUDED_TOPIC_PATTERNS = [re.compile(p, re.IGNORECASE) for p in [
    r"\bgender\b", r"\btransgender\b", r"\bnonbinary\b", r"\bnon-binary\b",
    r"\blgbtq?\b", r"\bqueer\b", r"\bsex\b", r"\bsexual\w*\b", r"\bheterosexual\b",
    r"\bhomosexual\b", r"\bgay\b", r"\blesbian\b", r"\bbisexual\b",
    r"\breligio\w*\b", r"\bchristian\w*\b", r"\bcatholic\w*\b", r"\bmuslim\w*\b",
    r"\bislam\w*\b", r"\bjewish\b", r"\bjudaism\b", r"\bhindu\w*\b", r"\bbuddhis\w*\b",
    r"\batheis\w*\b", r"\btheolog\w*\b", r"\bscripture\b", r"\bbible\b", r"\bquran\b",
    r"\bchurch\w*\b", r"\bfaith-based\b", r"\bevangelical\w*\b",
]]


def is_excluded_topic(row):
    blob = " ".join(str(cell) for cell in row if cell)
    return any(p.search(blob) for p in EXCLUDED_TOPIC_PATTERNS)

COLUMNS = [
    "title",
    "category",
    "topic",
    "primary_subject",
    "secondary_subject",
    "summary",
    "keywords",
    "source",
    "evidence_level",
    "source_url",
]


def sql_str(value):
    if value is None:
        return "null"
    text = str(value).strip()
    if not text:
        return "null"
    return "'" + text.replace("'", "''") + "'"


def sql_array(value):
    if not value:
        return "array[]::text[]"
    parts = [p.strip() for p in str(value).split(",") if p.strip()]
    if not parts:
        return "array[]::text[]"
    return "array[" + ",".join(sql_str(p) for p in parts) + "]"


def main():
    if len(sys.argv) != 3:
        print(f"Usage: {sys.argv[0]} <input.xlsx> <output.sql>", file=sys.stderr)
        sys.exit(1)

    in_path, out_path = sys.argv[1], sys.argv[2]

    wb = load_workbook(in_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(min_row=2, values_only=True)

    total = 0
    skipped_missing = 0
    skipped_topic = 0
    values_rows = []

    for row in rows:
        if row is None or all(c is None for c in row):
            continue
        total += 1
        (
            title,
            category,
            topic,
            primary_subject,
            secondary_subject,
            summary,
            keywords,
            source,
            evidence_level,
            source_url,
        ) = (row + (None,) * 10)[:10]

        if not title or not category or not summary or not primary_subject or not source or not evidence_level:
            skipped_missing += 1
            continue

        if is_excluded_topic(row):
            skipped_topic += 1
            continue

        values_rows.append(
            "("
            + ",".join(
                [
                    sql_str(title),
                    sql_str(category),
                    sql_str(topic or category),
                    sql_str(primary_subject),
                    sql_str(secondary_subject),
                    sql_str(summary),
                    sql_array(keywords),
                    sql_str(source),
                    sql_str(evidence_level),
                    sql_str(source_url),
                ]
            )
            + ")"
        )

    with open(out_path, "w", encoding="utf-8") as f:
        f.write("begin;\n\n")
        for i in range(0, len(values_rows), CHUNK_SIZE):
            chunk = values_rows[i : i + CHUNK_SIZE]
            f.write(
                f"insert into public.research_citations ({', '.join(COLUMNS)}) values\n"
            )
            f.write(",\n".join(chunk))
            f.write(";\n\n")
        f.write("commit;\n")

    print("=== research_citations conversion report ===", file=sys.stderr)
    print(f"Rows read:            {total}", file=sys.stderr)
    print(f"Skipped (missing required field): {skipped_missing}", file=sys.stderr)
    print(f"Skipped (gender/sex/sexuality/religion topic): {skipped_topic}", file=sys.stderr)
    print(f"Rows written:         {len(values_rows)}", file=sys.stderr)
    print(f"Output:               {out_path}", file=sys.stderr)


if __name__ == "__main__":
    main()
